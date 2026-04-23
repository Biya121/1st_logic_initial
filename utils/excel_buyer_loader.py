"""엑셀 바이어 데이터 로더 — SG_omethyl_omega3_2g / SG_sereterol_activair 전용.

엑셀 구조:
  Row 0-2: 제목·설명 (skip)
  Row 3:   컬럼 헤더
  Row 4+:  데이터 (각 20개)

반환 형식: buyer_scorer.rank_companies()가 바로 소비할 수 있는 list[dict].
  {
    company_name, country, website, contact, location,
    ingredient_match, source_region,
    enriched: { revenue, has_gmp, import_history, has_pharmacy_chain,
                public_channel, private_channel, has_target_country_presence,
                recommendation_reason, company_overview_kr,
                mah_capable(null), korea_experience("-"), ... }
  }
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any

_SHEET_MAP: dict[str, str] = {
    "SG_omethyl_omega3_2g":  "① Omethyl Cutielet (Omega-3)",
    "SG_sereterol_activair": "② Sereterol Activair (Inhaler)",
}

_INGREDIENT_KEYWORDS: dict[str, list[str]] = {
    "SG_omethyl_omega3_2g": [
        "omega-3", "omega3", "epa", "dha", "fish oil",
        "에틸에스테르", "ethyl ester", "lovaza", "omacor", "오메가",
    ],
    "SG_sereterol_activair": [
        "fluticasone", "salmeterol", "ics", "laba",
        "seretide", "advair", "inhaler", "흡입", "플루티카손", "살메테롤",
    ],
}

_PUBLIC_CHANNEL_KEYWORDS  = ["병원", "조달", "공공", "ndf", "gebiz", "hospital", "procurement", "government"]
_PRIVATE_CHANNEL_KEYWORDS = ["약국", "도매", "유통", "distributor", "wholesale", "pharmacy", "retail"]


def _excel_path() -> Path | None:
    env = os.environ.get("EXCEL_BUYER_PATH", "").strip()
    if env:
        p = Path(env)
        return p if p.is_file() else None
    candidates = [
        Path.home() / "Downloads" / "싱가포르_의약품_바이어_Omethyl_Sereterol_각20개.xlsx",
        Path(__file__).resolve().parents[1] / "datas" / "sg_buyers.xlsx",
    ]
    for c in candidates:
        if c.is_file():
            return c
    return None


def _bool_from_cell(val: Any) -> bool | None:
    s = str(val or "").strip()
    if "✅" in s or s.lower() in ("true", "yes", "o"):
        return True
    if "❌" in s or s.lower() in ("false", "no", "x"):
        return False
    return None


def _ingredient_match(pipeline_text: str, product_key: str) -> bool:
    t = (pipeline_text or "").lower()
    return any(kw in t for kw in _INGREDIENT_KEYWORDS.get(product_key, []))


def _infer_channels(company_type: str) -> tuple[bool | None, bool | None]:
    t = (company_type or "").lower()
    pub  = True if any(k in t for k in _PUBLIC_CHANNEL_KEYWORDS)  else None
    priv = True if any(k in t for k in _PRIVATE_CHANNEL_KEYWORDS) else None
    return pub, priv


def _parse_website(contact: str) -> str:
    m = re.search(r"(?:web|site|url)[:\s]*([^\s\n]+)", contact or "", re.IGNORECASE)
    if m:
        return m.group(1).strip().rstrip(",")
    m2 = re.search(r"([\w.-]+\.(com|sg|net|org|io)[\w/.-]*)", contact or "", re.IGNORECASE)
    if m2:
        return m2.group(1).strip()
    return "-"


def load_buyers(product_key: str) -> list[dict[str, Any]]:
    """엑셀에서 product_key에 해당하는 바이어 목록을 로드."""
    sheet_name = _SHEET_MAP.get(product_key)
    if not sheet_name:
        return []

    path = _excel_path()
    if not path:
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception:
        return []

    _SHEET_IDX = {"SG_omethyl_omega3_2g": 0, "SG_sereterol_activair": 1}
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        idx = _SHEET_IDX.get(product_key)
        if idx is None or idx >= len(wb.worksheets):
            return []
        ws = wb.worksheets[idx]

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 5:
        return []

    # 첫 번째 컬럼이 숫자인 행을 데이터 시작점으로 동적 탐지
    data_start = next(
        (i for i, r in enumerate(rows) if r and isinstance(r[0], (int, float))),
        4,
    )

    buyers: list[dict[str, Any]] = []

    for row in rows[data_start:]:
        if not row or row[0] is None:
            continue

        no_val       = row[0]
        name         = str(row[1] or "").strip()
        company_type = str(row[2] or "").strip()
        location     = str(row[3] or "").strip()
        contact      = str(row[4] or "").strip()
        revenue      = str(row[5] or "").strip()
        pipeline     = str(row[6] or "").strip()
        factory_cell = row[7]
        import_cell  = row[8]
        pharmacy_cell= row[9]
        note         = str(row[10] or "").strip()

        if not name:
            continue

        has_gmp          = _bool_from_cell(factory_cell)
        import_history   = _bool_from_cell(import_cell)
        has_pharmacy     = _bool_from_cell(pharmacy_cell)
        pub_ch, priv_ch  = _infer_channels(company_type)
        ing_match        = _ingredient_match(pipeline, product_key)
        website          = _parse_website(contact)

        # 소재지에 Singapore 명시 → 싱가포르 진출 확인
        sg_presence: bool | None = True if "singapore" in location.lower() else None

        enriched: dict[str, Any] = {
            "revenue":                    revenue or "-",
            "employees":                  "-",
            "founded":                    "-",
            "territories":                ["Singapore"] if sg_presence else [],
            "has_target_country_presence": sg_presence,
            "has_gmp":                    has_gmp,
            "import_history":             import_history,
            "procurement_history":        None,
            "has_pharmacy_chain":         has_pharmacy,
            "public_channel":             pub_ch,
            "private_channel":            priv_ch,
            "mah_capable":               None,       # enricher가 보완
            "korea_experience":          "-",         # enricher가 보완
            "certifications":            [],
            "source_urls":               [website] if website != "-" else [],
            "company_overview_kr":       "-",         # enricher가 보완
            "recommendation_reason":     note or "-",
            "_pipeline_text":            pipeline,    # enricher 참조용 (내부 전용)
            "_company_type":             company_type,
        }

        buyers.append({
            "company_name":       name,
            "country":            "Singapore",
            "website":            website,
            "contact":            contact,
            "location":           location,
            "company_type":       company_type,
            "ingredient_match":   ing_match,
            "source_region":      "excel",
            "enriched":           enriched,
        })

    return buyers
